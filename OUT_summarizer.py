import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

st.title("📊 OUT Summarizer — Country Report Generator")

uploaded_file = st.file_uploader("📂 Upload Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)
    df.columns = df.columns.str.strip()

    # Fill merged cells
    for col in ["Start", "End", "Channel", "ID"]:
        if col in df.columns:
            df[col] = df[col].ffill()

    if 'Country' not in df.columns:
        st.error("❌ Column 'Country' not found!")
        st.stop()
    df['Country'] = df['Country'].astype(str).str.strip().str.upper()

    # =========================
    # PARSOWANIE DAT START
    # =========================
    if 'Start' not in df.columns:
        st.error("❌ Column 'Start' missing!")
        st.stop()

    # YY/MM/DD -> YYYY-MM-DD
    df['Start'] = pd.to_datetime(
        df['Start'].astype(str).str.strip(),
        format="%y/%m/%d",
        errors="coerce"
    ).dt.date

    parsed = df['Start'].notna().sum()
    if parsed == 0:
        st.error("❌ No valid dates in 'Start'. Check Excel formatting.")
        st.stop()

    min_date = df['Start'].min()
    max_date = df['Start'].max()

    date_range = st.date_input(
        "📅 Select date range (based on 'Start')",
        value=(min_date, max_date),
        min_value=min_date,
        max_value=max_date
    )

    # =========================
    # CATEGORY
    # =========================
    if 'Category' not in df.columns:
        st.error("❌ Column 'Category' missing!")
        st.stop()

    df['Category_norm'] = df['Category'].astype(str).str.strip().str.lower()
    categories = sorted(df['Category_norm'].unique())
    categories_with_all = ["All"] + categories

    selected_category = st.selectbox(
        "🏷️ Select category",
        categories_with_all,
        index=0
    )

    # =========================
    # GENERATE REPORT
    # =========================
    if st.button("🚀 Generate report"):
        start_date, end_date = date_range

        if selected_category == "All":
            mask = (df['Start'] >= start_date) & (df['Start'] <= end_date)
        else:
            mask = (
                (df['Category_norm'] == selected_category) &
                (df['Start'] >= start_date) &
                (df['Start'] <= end_date)
            )

        filtered_df = df.loc[mask].copy()
        if filtered_df.empty:
            st.warning("⚠️ No data for these filters.")
            st.stop()

        countries = sorted(filtered_df["Country"].dropna().unique())

        keep_cols = [
            'Start', 'End', 'Channel', 'ID', 'Name', 'Description', 'Category',
            'Visits', 'Orders', 'Demand', 'CVR', 'AOV',
            'Expected Demand', 'Demand Diff to Expected', '% Expected Demand',
            'Country'
        ]
        existing_cols = [c for c in keep_cols if c in filtered_df.columns]
        filtered_df = filtered_df[existing_cols]

        numeric_cols = [
            'Visits', 'Orders', 'Demand', 'CVR', 'AOV',
            'Expected Demand', 'Demand Diff to Expected', '% Expected Demand'
        ]
        numeric_cols = [c for c in numeric_cols if c in filtered_df.columns]

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for country in countries:
                country_df = filtered_df[filtered_df['Country'] == country].copy()
                if not country_df.empty:

                    if "Demand" in country_df.columns:
                        country_df = country_df.sort_values(by="Demand", ascending=False)

                    means = country_df[numeric_cols].mean().to_frame().T
                    means.index = ['Average']

                    if 'Orders' in means.columns:
                        means['Orders'] = round(means['Orders'])

                    for col in ['Demand', 'Expected Demand']:
                        if col in means.columns:
                            means[col] = round(means[col], 2)

                    for col in ['CVR', '% Expected Demand']:
                        if col in means.columns:
                            means[col] = round(means[col], 4)

                    country_df[numeric_cols] = country_df[numeric_cols].round(2)

                    final_df = pd.concat([country_df, means], ignore_index=False)
                    final_df.to_excel(writer, index=False, sheet_name=str(country))

                    ws = writer.sheets[str(country)]

                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)

                    yellow_fill = PatternFill("solid", fgColor="FFFF00")
                    red_fill = PatternFill("solid", fgColor="FF9999")
                    green_fill = PatternFill("solid", fgColor="99FF99")
                    light_blue_fill = PatternFill("solid", fgColor="ADD8E6")

                    # Headers
                    for col_idx, col in enumerate(final_df.columns, 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = max(
                            12,
                            max(
                                (len(str(cell.value)) if cell.value is not None else 0
                                 for cell in ws[get_column_letter(col_idx)]),
                                default=12
                            )
                        )
                        cell = ws.cell(row=1, column=col_idx)
                        cell.fill = yellow_fill
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border

                    # Data + summary rows
                    for row_idx in range(2, ws.max_row + 1):
                        for col_idx, col in enumerate(final_df.columns, 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.border = border
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                            # FORMAT DAT
                            if col in ["Start", "End"]:
                                cell.number_format = "yyyy-mm-dd"

                            if row_idx < ws.max_row:
                                if col in ["Demand", "Expected Demand", "Demand Diff to Expected"]:
                                    cell.number_format = '€#,##0.00'
                                    if col in ["Demand Diff to Expected", "% Expected Demand"] and isinstance(cell.value, (int, float)):
                                        cell.fill = green_fill if cell.value >= 0 else red_fill

                                if col in ["CVR", "% Expected Demand"]:
                                    cell.number_format = '0.00%'
                            else:
                                cell.fill = light_blue_fill
                                if col == "Category":
                                    cell.value = "ARV"

                                if col in ["Demand", "Expected Demand", "Demand Diff to Expected"]:
                                    cell.number_format = '€#,##0.00'
                                if col in ["CVR", "% Expected Demand"]:
                                    cell.number_format = '0.00%'

            for empty_sheet in ["Brands", "Category", "Stock level", "Conclusions"]:
                pd.DataFrame().to_excel(writer, sheet_name=empty_sheet)

        output.seek(0)
        filename = f"{selected_category}_{start_date}_{end_date}.xlsx".replace(" ", "_")

        st.success("✅ Report ready!")
        st.download_button(
            "📥 Download report",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument-spreadsheetml.sheet"
        )

